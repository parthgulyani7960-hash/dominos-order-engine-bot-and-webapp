import os
import csv
import re
import shutil
import datetime
import time
import logging

logger = logging.getLogger(__name__)
import html
from collections import defaultdict
from cryptography.fernet import Fernet

def escape_html(text: str) -> str:
    """Safely escapes HTML tags and special characters for Telegram HTML parse mode."""
    if not text:
        return ""
    return html.escape(str(text))

# Setup key path
DATA_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "data"))
BACKUP_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "backups"))
KEY_FILE = os.path.join(DATA_DIR, "secret.key")
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(BACKUP_DIR, exist_ok=True)

# --- Symmetric Encryption / Decryption ---

def get_encryption_key():
    """Gets or generates a persistent key for encryption."""
    if os.path.exists(KEY_FILE):
        with open(KEY_FILE, "rb") as f:
            return f.read()
    else:
        key = Fernet.generate_key()
        with open(KEY_FILE, "wb") as f:
            f.write(key)
        return key

_fernet = Fernet(get_encryption_key())

def encrypt_data(data: str) -> str:
    """Encrypts a string and returns a base64 string representation."""
    if not data:
        return ""
    return _fernet.encrypt(data.encode()).decode()

def decrypt_data(token: str) -> str:
    """Decrypts a base64 string representation back to string."""
    if not token:
        return ""
    try:
        return _fernet.decrypt(token.encode()).decode()
    except Exception:
        return "[Decryption Failed]"

# --- Database Backups ---

def run_backup():
    """Copies the active sqlite database to the backups folder and cleans up old backups."""
    db_path = os.path.join(DATA_DIR, "pizza.db")
    if not os.path.exists(db_path):
        return None
    
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_file = os.path.join(BACKUP_DIR, f"pizza_backup_{timestamp}.db")
    
    try:
        shutil.copy2(db_path, backup_file)
        
        # Cleanup: Keep only the 7 most recent backups
        backups = sorted(
            [os.path.join(BACKUP_DIR, f) for f in os.listdir(BACKUP_DIR) if f.startswith("pizza_backup_")],
            key=os.path.getmtime
        )
        while len(backups) > 7:
            old_backup = backups.pop(0)
            os.remove(old_backup)
            
        return backup_file
    except Exception as e:
        logger.error(f"Backup failed: {e}")
        return None

# --- Gift Card File Parsing ---

def parse_gift_card_file(file_path: str, filename: str):
    """
    Parses a file (CSV, XLSX, PDF) containing gift card records.
    Returns a list of dicts: [{'code': str, 'pin': str, 'value': float}]
    """
    ext = os.path.splitext(filename.lower())[1]
    cards = []
    
    if ext == ".csv":
        with open(file_path, "r", encoding="utf-8") as f:
            # Handle delimiter variations (comma, semicolon, tab)
            content = f.read(2048)
            f.seek(0)
            delimiter = ","
            if ";" in content and content.count(";") > content.count(","):
                delimiter = ";"
            elif "\t" in content:
                delimiter = "\t"
                
            reader = csv.reader(f, delimiter=delimiter)
            header = [h.strip().lower() for h in next(reader, [])]
            
            # Map column headers to indices
            code_idx, pin_idx, val_idx = -1, -1, -1
            for i, h in enumerate(header):
                if "code" in h or "card" in h:
                    code_idx = i
                elif "pin" in h:
                    pin_idx = i
                elif "value" in h or "amount" in h:
                    val_idx = i
            
            # Fallback to column positions if headers are not recognized
            if code_idx == -1: code_idx = 0
            if pin_idx == -1: pin_idx = 1
            if val_idx == -1: val_idx = 2
            
            for row in reader:
                if not row or len(row) < 2:
                    continue
                code = row[code_idx].strip() if len(row) > code_idx else ""
                pin = row[pin_idx].strip() if len(row) > pin_idx else ""
                val = 100.0
                if val_idx != -1 and len(row) > val_idx:
                    try:
                        val = float(row[val_idx].replace("$", "").replace("₹", "").replace(",", "").strip())
                    except ValueError:
                        val = 100.0
                
                if code and pin and val > 0:
                    cards.append({"code": code, "pin": pin, "value": val})
                    
    elif ext in [".xlsx", ".xls"]:
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ValueError("openpyxl library is required to parse Excel files.")
        wb = load_workbook(file_path, read_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = [str(cell).strip().lower() if cell else "" for cell in next(rows, [])]
        
        code_idx, pin_idx, val_idx = -1, -1, -1
        for i, h in enumerate(header):
            if "code" in h or "card" in h:
                code_idx = i
            elif "pin" in h:
                pin_idx = i
            elif "value" in h or "amount" in h:
                val_idx = i
                
        if code_idx == -1: code_idx = 0
        if pin_idx == -1: pin_idx = 1
        if val_idx == -1: val_idx = 2
        
        for row in rows:
            if not row or len(row) <= max(code_idx, pin_idx, val_idx):
                continue
            code = str(row[code_idx]).strip() if row[code_idx] is not None else ""
            pin = str(row[pin_idx]).strip() if row[pin_idx] is not None else ""
            try:
                val = float(str(row[val_idx]).replace("$", "").replace("₹", "").replace(",", "").strip()) if row[val_idx] is not None else 0.0
            except ValueError:
                continue
                
            if code and pin and val > 0:
                cards.append({"code": code, "pin": pin, "value": val})
                
    elif ext == ".pdf":
        try:
            from pypdf import PdfReader
        except ImportError:
            raise ValueError("pypdf library is required to parse PDF files.")
        reader = PdfReader(file_path)
        full_text = ""
        for page in reader.pages:
            text = page.extract_text()
            if text:
                full_text += text + "\n"
        
        # Regex patterns to parse gift card info
        # Looks for combinations like: Card Code: XXXX, PIN: YYYY, Value: ZZZ
        # Or: XXXX-XXXX YYYY 50.00
        patterns = [
            r"(?:code|card|gift\s*card)?[:\s\-]*([A-Z0-9\-]{8,25})\b.*?(?:pin)[:\s]*(\d{4,8})\b.*?([0-9\.]+)",
            r"\b([A-Z0-9\-]{8,25})\s+(\d{4,8})\s+\$?([0-9\.]+)\b"
        ]
        
        lines = full_text.split("\n")
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            matched = False
            for pat in patterns:
                m = re.search(pat, line, re.IGNORECASE)
                if m:
                    code = m.group(1).strip()
                    pin = m.group(2).strip()
                    try:
                        val = float(m.group(3).strip())
                    except ValueError:
                        continue
                    
                    if code and pin and val > 0:
                        cards.append({"code": code, "pin": pin, "value": val})
                        matched = True
                        break
            if not matched:
                # Try simple regex: Card code (alphanumeric + hyphen), Pin (4-8 digits), Value (number)
                # This catches columns or spaces in tables
                parts = re.findall(r"\b([A-Z0-9\-]{8,25})\b|\b(\d{4,8})\b|\b([0-9\.]+)\b", line, re.IGNORECASE)
                # parts is list of tuples. Filter empty matches
                flat_parts = []
                for pt in parts:
                    for val in pt:
                        if val: flat_parts.append(val)
                if len(flat_parts) >= 3:
                    # Let's see if we can identify code, pin and value
                    code_cand, pin_cand, val_cand = "", "", 0.0
                    for fp in flat_parts:
                        if fp.isdigit() and len(fp) >= 4 and len(fp) <= 8 and not pin_cand:
                            pin_cand = fp
                        elif re.match(r"^[A-Z0-9\-]{8,25}$", fp, re.IGNORECASE) and not code_cand:
                            code_cand = fp
                        else:
                            try:
                                val_cand = float(fp.replace("$", "").replace("₹", ""))
                            except ValueError:
                                pass
                    if code_cand and pin_cand and val_cand > 0:
                        cards.append({"code": code_cand, "pin": pin_cand, "value": val_cand})
                        
    return cards

# --- Sliding Window Rate Limiter ---

class RateLimiter:
    def __init__(self, limit: int = 60, window: int = 60):
        """
        limit: Max requests allowed
        window: Sliding window in seconds
        """
        self.limit = limit
        self.window = window
        self.history = defaultdict(list)

    def is_rate_limited(self, key: str) -> bool:
        now = time.time()
        # Filter request timestamps older than sliding window
        self.history[key] = [t for t in self.history[key] if now - t < self.window]
        
        if len(self.history[key]) >= self.limit:
            return True
        
        self.history[key].append(now)
        return False

# Global instance for API routes
api_rate_limiter = RateLimiter(limit=100, window=60)
# Tighter rate limit for wallet & login routes
strict_rate_limiter = RateLimiter(limit=10, window=60)

# --- NPCI Compliant UPI URI & QR Generator ---

import urllib.parse
import io
import base64
import qrcode

def generate_upi_qr_details(upi_id: str, upi_name: str, amount: float, ref_id: str, note: str = "") -> dict:
    """
    Generates a 100% NPCI (National Payments Corporation of India) compliant UPI URI string
    and high-resolution QR code images (both base64 data URI and high-res HTTP API fallback URL).

    Returns dict with:
      - upi_uri: NPCI compliant UPI scheme string (encoded with %20, cu=INR, am formatted to 2 decimals)
      - qr_data_url: High resolution base64 PNG data URI with 4-module quiet zone & medium ECC
      - qr_code_url: High quality external API URL fallback with 4-module quiet zone & ECC
    """
    clean_id = (upi_id or "pranjalnautry@fam").strip()
    clean_name = (upi_name or "Domino's Order Engine").strip()
    clean_ref = (ref_id or "").strip()
    clean_note = (note or f"Order {clean_ref}").strip()
    
    # URL encode parameters with %20 (NPCI specification requires %20, NEVER '+')
    enc_name = urllib.parse.quote(clean_name, safe='')
    enc_note = urllib.parse.quote(clean_note, safe='')
    amt_str = f"{float(amount):.2f}"
    
    # Construct NPCI Standard UPI Deep Link URI
    upi_uri = f"upi://pay?pa={clean_id}&pn={enc_name}&am={amt_str}&cu=INR&tr={clean_ref}&tn={enc_note}"
    
    # 1. Generate local high-res 400x400 PNG with 4-module white quiet zone border & medium ECC
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=10,
        border=4,  # Standard quiet zone margin required for phone cameras
    )
    qr.add_data(upi_uri)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    
    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    b64_str = base64.b64encode(buffer.getvalue()).decode("utf-8")
    qr_data_url = f"data:image/png;base64,{b64_str}"
    
    # 2. External API fallback URL with quiet zone margin (margin=4) and high error correction (ecc=M)
    encoded_uri_for_api = urllib.parse.quote(upi_uri, safe='')
    qr_code_url = f"https://api.qrserver.com/v1/create-qr-code/?size=400x400&ecc=M&margin=4&data={encoded_uri_for_api}"
    
    return {
        "upi_id": clean_id,
        "upi_name": clean_name,
        "upi_uri": upi_uri,
        "qr_data_url": qr_data_url,
        "qr_code_url": qr_code_url
    }
